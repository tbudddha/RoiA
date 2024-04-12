import openpyxl

def dataTransfer(inpath, outpath):
    outpath = 'EXAMPLE_OUTPUT_FILENAME.xlsx'
    myWorkbook = openpyxl.load_workbook(filename=outpath, keep_vba=False)

    inpath = 'EXAMPLE_INPUT_FILENAME.xlsx'
    orig = openpyxl.load_workbook(filename=inpath, keep_vba=False)

    dapiNameCount = 2
    CHANNEL1NameCount = 2
    CHANNEL2NameCount = 2

    for x in range(1, LAST-COLUMN NUMBER, 20):
        category_edit = orig['Sheet1'].cell(row=1, column=x).value

        print(category_edit)
        if (category_edit == ""):
            print("EXTRA SPACE")
            exit()

        myWorkbook['DAPI_FIRST'].cell(row=dapiNameCount, column=1, value=category_edit)
        myWorkbook['DAPI_SECOND'].cell(row=dapiNameCount, column=1, value=category_edit)
        dapiNameCount += 1
        myWorkbook['CHANNEL1_FIRST'].cell(row=CHANNEL1NameCount, column=1, value=category_edit)
        myWorkbook['CHANNEL1_SECOND'].cell(row=CHANNEL1NameCount, column=1, value=category_edit)
        myWorkbook['PA_CHANNEL1_FIRST'].cell(row=CHANNEL1NameCount, column=1, value=category_edit)
        myWorkbook['PA_CHANNEL1_SECOND'].cell(row=CHANNEL1NameCount, column=1, value=category_edit)
        myWorkbook['CHANNEL1_vs_CHANNEL2_FIRST'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        myWorkbook['CHANNEL1_vs_CHANNEL2_SECOND'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        CHANNEL1NameCount += 1
        myWorkbook['CHANNEL2_FIRST'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        myWorkbook['CHANNEL2_SECOND'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        myWorkbook['PA_CHANNEL2_FIRST'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        myWorkbook['PA_CHANNEL2_SECOND'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        myWorkbook['Stats_CHANNEL1_vs_CHANNEL2_FIRST'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        myWorkbook['Stats_CHANNEL1_vs_CHANNEL2_SECOND'].cell(row=CHANNEL2NameCount, column=1, value=category_edit)
        CHANNEL2NameCount += 1

    myWorkbook.save(outpath)

    # Add channel measurements to appropriate channel sheets
    dapiCount = 2
    CHANNEL1NameCount = 2
    CHANNEL2NameCount = 2
    DAPI = []
    chan1 = []
    chan2 = []

    for x in range(1, LAST-COLUMN NUMBER, 20):
        for y in range(x, x + 19, 1):
            category_edit = orig['Sheet1'].cell(row=3, column=y).value
            DAPI.append(category_edit)
            myWorkbook['DAPI_FIRST'].cell(row=dapiCount, column=((y % 20) + 1), value=category_edit)
            category_edit = orig['Sheet1'].cell(row=4, column=y).value
            DAPI.append(category_edit)
            myWorkbook['DAPI_SECOND'].cell(row=dapiCount, column=((y % 20) + 1), value=category_edit)
        dapiCount += 1
        print(DAPI)
        DAPI = []

        for i in range(x, x + 19, 1):
            category_edit = orig['Sheet1'].cell(row=5, column=i).value
            chan1.append(category_edit)
            myWorkbook['CHANNEL1_FIRST'].cell(row=CHANNEL1NameCount, column=((i % 20) + 1), value=category_edit)
            category_edit = orig['Sheet1'].cell(row=6, column=i).value
            chan1.append(category_edit)
            myWorkbook['CHANNEL1_SECOND'].cell(row=CHANNEL1NameCount, column=((i % 20) + 1), value=category_edit)
        CHANNEL1NameCount += 1
        print(chan1)
        chan1 = []

        for j in range(x, x + 19, 1):
            category_edit = orig['Sheet1'].cell(row=7, column=j).value
            chan2.append(category_edit)
            myWorkbook['CHANNEL2_FIRST'].cell(row=CHANNEL2NameCount, column=((j % 20) + 1), value=category_edit)
            category_edit = orig['Sheet1'].cell(row=8, column=j).value
            chan2.append(category_edit)
            myWorkbook['CHANNEL2_SECOND'].cell(row=CHANNEL2NameCount, column=((j % 20) + 1), value=category_edit)
        CHANNEL2NameCount += 1
        print(chan2)
        chan2 = []
    myWorkbook.save(outpath)

    CHANNEL1NameCount = 2
    CHANNEL2NameCount = 2
    #DAPI = []
    chan1 = []
    chan2 = []

    for x in range(1, LAST-COLUMN NUMBER, 20):
        for i in range(x, x + 19, 1):
            category_edit = orig['Sheet1'].cell(row=9, column=i).value
            chan1.append(category_edit)
            myWorkbook['PA_CHANNEL1_FIRST'].cell(row=CHANNEL1NameCount, column=((i % 20) + 1), value=category_edit)
            category_edit = orig['Sheet1'].cell(row=11, column=i).value
            chan1.append(category_edit)
            myWorkbook['PA_CHANNEL1_SECOND'].cell(row=CHANNEL1NameCount, column=((i % 20) + 1), value=category_edit)
        CHANNEL1NameCount += 1
        print(chan1)
        chan1 = []

        for j in range(x, x + 19, 1):
            category_edit = orig['Sheet1'].cell(row=10, column=j).value
            chan2.append(category_edit)
            myWorkbook['PA_CHANNEL2_FIRST'].cell(row=CHANNEL2NameCount, column=((j % 20) + 1), value=category_edit)
            category_edit = orig['Sheet1'].cell(row=12, column=j).value
            chan2.append(category_edit)
            myWorkbook['PA_CHANNEL2_SECOND'].cell(row=CHANNEL2NameCount, column=((j % 20) + 1), value=category_edit)
        CHANNEL2NameCount += 1
        print(chan2)
        chan2 = []
    myWorkbook.save(outpath)

    colocCount = 2
    statsCount = 2
    COLOC = []
    STATS = []

    for x in range(1, LAST-COLUMN NUMBER, 20):
        for y in range(x, x + 19, 1):
            category_edit = orig['Sheet1'].cell(row=13, column=y).value
            COLOC.append(category_edit)
            myWorkbook['CHANNEL1_vs_CHANNEL2_FIRST'].cell(row=colocCount, column=((y % 20) + 1), value=category_edit)
            category_edit = orig['Sheet1'].cell(row=14, column=y).value
            COLOC.append(category_edit)
            myWorkbook['CHANNEL1_vs_CHANNEL2_SECOND'].cell(row=colocCount, column=((y % 20) + 1), value=category_edit)
        colocCount += 1
        print(COLOC)
        COLOC = []

        for i in range(x, x + 19, 1):
            category_edit = orig['Sheet1'].cell(row=15, column=i).value
            STATS.append(category_edit)
            myWorkbook['Stats_CHANNEL1_vs_CHANNEL2_FIRST'].cell(row=statsCount, column=((i % 20) + 1), value=category_edit)
            category_edit = orig['Sheet1'].cell(row=16, column=i).value
            STATS.append(category_edit)
            myWorkbook['Stats_CHANNEL1_vs_CHANNEL2_SECOND'].cell(row=statsCount, column=((i % 20) + 1), value=category_edit)
        statsCount += 1
        print(STATS)
        STATS = []
    myWorkbook.save(outpath)


if __name__ == "__main__":
    dataTransfer("EXAMPLE_INPUT_FILENAME.xlsx","EXAMPLE_OUTPUT_FILENAME.xlsx")
